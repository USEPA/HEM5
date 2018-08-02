# -*- coding: utf-8 -*-
"""
Created on Mon Oct 23 12:43:52 2017

@author: dlindsey
"""
import inspect
import math
import time
import numpy as np
import openpyxl
import pandas as pd

#from writer.csv.AllPolarReceptors import AllPolarReceptors
#from writer.csv.CsvWriter import CsvWriter
#from writer.excel.AcuteBreakdown import AcuteBreakdown
from writer.csv import *
import sys

class Process_outputs():
    
    def __init__(self, outdir, facid, model, prep, runstream):
        self.facid = facid
        self.haplib_m = model.haplib.dataframe.as_matrix()
        self.hapemis = runstream.hapemis
        self.outdir = outdir
        self.inner_m = prep.innerblks.as_matrix()
        self.outerblks_df = prep.outerblks
        self.polar_recs = runstream.polar_df
        self.numsectors = self.polar_recs["sector"].max()
        self.numrings = self.polar_recs["ring"].max()
        self.model = model

        self.model.runstream_polar_recs = runstream.polar_df
        self.model.runstream_hapemis = runstream.hapemis

        # Units conversion factor
        self.cf = 2000*0.4536/3600/8760
#        #first check for facilities folder, if not create as output directory
#        if not os.path.exists(self.outdir):
#            os.makedirs(self.outdir)
        
    
    def create_all_polar_receptors(self, polarplot_df):
        
        """
        
        Create the facility specific All Polar Receptor output file.
        
        This is a CSV formatted file with the following fields:
            source_id
            emis_type
            pollutant
            conc_ug_m3
            distance (m)
            angle
            sector number
            ring number
            elevation (m)
            latitude
            longitude
            overlap (Y/N)
            
        """
             
        # array of unique source_id's
        srcids = polarplot_df['source_id'].unique().tolist()
    
        # process polar concs one source_id at a time
        for x in srcids:
            polarplot_onesrcid = polarplot_df[["utme","utmn","source_id","result"]].loc[polarplot_df['source_id'] == x]
            hapemis_onesrcid = self.hapemis[["source_id","pollutant","emis_tpy"]].loc[self.hapemis['source_id'] == x]
            collist = ["source_id", "emis_type", "pollutant", "conc_ug_m3", "distance", "angle", "sector", "ring_no", "elev", "lat", "lon", "overlap"]
            dlist = []
            for row1 in polarplot_onesrcid.itertuples():
                for row2 in hapemis_onesrcid.itertuples():
                    d_sourceid = row1[3]
                    d_emistype = "C"
                    d_pollutant = row2[2]
                    d_conc = row1[4] * row2[3] * self.cf
                    d_distance = self.polar_recs.loc[(self.polar_recs["utme"] == row1[1]) & (self.polar_recs["utmn"] == row1[2]), "distance"].values[0]
                    d_angle = self.polar_recs.loc[(self.polar_recs["utme"] == row1[1]) & (self.polar_recs["utmn"] == row1[2]), "angle"].values[0]
                    d_sector = self.polar_recs.loc[(self.polar_recs["utme"] == row1[1]) & (self.polar_recs["utmn"] == row1[2]), "sector"].values[0]
                    d_ring_no = self.polar_recs.loc[(self.polar_recs["utme"] == row1[1]) & (self.polar_recs["utmn"] == row1[2]), "ring"].values[0]
                    d_elev = self.polar_recs.loc[(self.polar_recs["utme"] == row1[1]) & (self.polar_recs["utmn"] == row1[2]), "elev"].values[0]
                    d_lat = self.polar_recs.loc[(self.polar_recs["utme"] == row1[1]) & (self.polar_recs["utmn"] == row1[2]), "lat"].values[0]
                    d_lon = self.polar_recs.loc[(self.polar_recs["utme"] == row1[1]) & (self.polar_recs["utmn"] == row1[2]), "lon"].values[0]
                    d_overlap = ""
                    datalist = [d_sourceid, d_emistype, d_pollutant, d_conc, d_distance, d_angle, d_sector, d_ring_no, d_elev, d_lat, d_lon, d_overlap]
                    dlist.append(dict(zip(collist, datalist)))

        polarconc_df = pd.DataFrame(dlist, columns=collist)
                
        return polarconc_df
    

                 
    def polar_build(self, r):
        

        found = np.logical_and(self.polar_m[:,3] == r[0], self.polar_m[:,4] == r[1])
    
        polgridmore = self.polar_m[found]
        
        step1 = np.append(r, polgridmore[0][1:3])
        step2 = np.append(step1, polgridmore[0][5:10])
        
        
        #print(step2.shape[0])
        
        #write to matrix
        self.polgridmore = np.append(self.polgridmore, [step2], axis = 0)
        
        # Compute pollutant specific air concentrations 
        polarc = self.hap_m[self.hap_m[:, 1] == r[7]]
        
        
        for p in polarc:
            
            
            polarconc1 = np.append(step2, p[0])
            polarconc2 = np.append(polarconc1, p[2:4])
           
        
            #do some math
            polarpolconc = np.insert(polarconc2, 20 ,(polarconc2[2] * polarconc2[19] * self.cf))
            
            #write to matrix
            self.polar_pol_con = np.append(self.polar_pol_con, [polarpolconc], axis = 0)
        
            #get  ure and rfc for each pollutant
            prisk = self.haplib_m[self.haplib_m[:,0] == polarpolconc[18]][0][3:5]
            
            prisk2 = np.append(polarpolconc, prisk)
            
            #add resp and neuro
            risktarg = self.target_m[self.target_m[:,0] == prisk2[18]][0][2:4]
            
            polarrisk = np.append(prisk2, risktarg)
            
            #multiply concentration by ure for risk
            risk = np.insert(polarrisk, 25, (polarrisk[20]*polarrisk[21] ))
            
            #calculate other risks
            resp_hi = np.insert(risk, 26, ((risk[23] * risk[20])/polarrisk[22]))
            neur_hi = np.insert(resp_hi, 27, ((risk[24]*risk[20])/polarrisk[22]))
        
            #write to matrix
            self.polar_risk= np.append(self.polar_risk, [neur_hi], axis = 0)
            
        

    def inner_build(self, row):
    
    
        #find utmn + utme matching coordinates
        found = np.logical_and(self.plot_m[:, 1] == row[10], self.plot_m[:,0] == row[11])
        plot = self.plot_m[found]
        
        blks = np.append(row[3:6], row[7])
        
        inner_1 = (np.append(plot, blks))
        #inner_1.append(self.plot_m[found])
        
        pol = self.hap_m[self.hap_m[:,1] == plot[0][7]]
        
        
        for p in pol:
        # indices for p
            
            sel1 = p[0]
            sel2 = p[2:4]
            
            sel = np.append(sel1, sel2)
            
            inner_conc = np.append(inner_1, sel)
            
            pol_conc = np.insert(inner_conc, 17, (inner_conc[2] * inner_conc[16]* self.cf))
            
            #write to pollutant concentration matrix
            self.pol_conc = np.append(self.pol_conc, [pol_conc], axis=0)
            
            #print(pol_conc)
            
            #now set up risk
            fp = self.haplib_m[self.haplib_m[:,0] == p[2]][0][3:5]
           
            #merge with pollutant concentration
            #in innerrisk matrix:
            # 0 = , 1 = , 2 =, 3 = , 4 = , 5 = , 6 = , 7 = , 8 = , 9 = , 10 = , 
            
            innerrisk0 = np.append(pol_conc, fp)
            
            
            #now get target endpoint values
            ft = self.target_m[self.target_m[:,0] == p[2]][0][2:]
            
            #merge with innerrisk 0 
            ## DO WE REALLY NEED TO APPEND THESE OR CAN WE USE THE VALUES IN FT TO CALCULATE THE "HI" VALUES FOR RESP ETC
            #innerrisk1 = np.append(innerrisk0, ft)
            
            # USE NP INSERT TO GET THE COMPUTED VALUES FOR HI
            # in ft matrix:
            # 0 = resp, 1 = liver, 2 = neuro, 3 = dev, 4 = reprod, 5 = kidney, 
            # 6 = ocular, 7 = endoc, 8 = hemato, 9 = immune, 10 = skeletal, 11 = spleen,
            # 12 = thyroid, 13 = wholebod
                   
            risk = np.insert(innerrisk0, 20, (innerrisk0[17] * innerrisk0[18]))
            resp = np.insert(risk, 21, ((ft[0] * innerrisk0[17])/innerrisk0[19]))
            liver = np.insert(resp, 22, ((ft[1] * innerrisk0[17])/innerrisk0[19]))
            neuro = np.insert(liver, 23, ((ft[2] * innerrisk0[17])/innerrisk0[19]))
            deve = np.insert(neuro, 24, ((ft[3] * innerrisk0[17])/innerrisk0[19]))
            reprod = np.insert(deve, 25, ((ft[4] * innerrisk0[17])/innerrisk0[19]))
            kidney = np.insert(reprod, 26, ((ft[5] * innerrisk0[17])/innerrisk0[19]))
            ocular = np.insert(kidney, 27, ((ft[6] * innerrisk0[17])/innerrisk0[19]))
            endoc = np.insert(ocular, 28, ((ft[7] * innerrisk0[17])/innerrisk0[19]))
            hemato = np.insert(endoc, 29, ((ft[8] * innerrisk0[17])/innerrisk0[19]))
            immune = np.insert(hemato, 30, ((ft[9] * innerrisk0[17])/innerrisk0[19]))
            skeletal = np.insert(immune, 31, ((ft[10] * innerrisk0[17])/innerrisk0[19]))
            spleen = np.insert(skeletal, 32, ((ft[11] * innerrisk0[17])/innerrisk0[19]))
            thyroid = np.insert(spleen, 33, ((ft[12] * innerrisk0[17])/innerrisk0[19]))
            wholebod = np.insert(thyroid, 34, ((ft[13] * innerrisk0[17])/innerrisk0[19]))
            
            #write as matrix
            self.finalrisk = np.append(self.finalrisk, [wholebod], axis = 0)
            

    def process(self):

        
        start = time.time()
    
        # Read plotfile and put into dataframe
        pfile = open("resources/plotfile.plt", "r")
        plot_df = pd.read_table(pfile, delim_whitespace=True, header=None, 
            names=['utme','utmn','result','elev','hill','flag','avg_time','source_id','num_yrs','net_id'],
            usecols=[0,1,2,3,4,5,6,7,8,9], 
            converters={'utme':np.float64,'utmn':np.float64,'result':np.float64,'elev':np.float64,'hill':np.float64
                   ,'flag':np.float64,'avg_time':np.str,'source_id':np.str,'num_yrs':np.int64,'net_id':np.str},
            comment='*')

        
        #----------- create All_Polar_Receptor output file -----------------
        all_polar_receptors = AllPolarReceptors(self.outdir, self.facid, self.model, plot_df)
        all_polar_receptors.write()
               
        
        #debug
        sys.exit()
        
        
        #convert to matrix called self.plot_m 
        self.plot_m = plot_df.as_matrix()
        
        #get polgrid 
        self.polgrid = polgrid_df.as_matrix()
        
      
        #get target organ endpoint dataframe
        targetendpt_df = pd.read_excel(r"resources/target_organ_endpoints.xlsx"
                 , names=("pollutant","epa_woe","resp","liver","neuro","dev","reprod","kidney","ocular","endoc"
                          ,"hemato","immune","skeletal","spleen","thyroid","wholebod")
                 , converters={"pollutant":str,"epa_woe":str,"resp":float,"liver":float,"neuro":float,"dev":float
                               ,"reprod":float,"kidney":float,"ocular":float,"endoc":float,"hemato":float,"immune":float
                               ,"skeletal":float,"spleen":float,"thyroid":float,"wholebod":float})

        
        #convert to matrix
        self.target_m = targetendpt_df.as_matrix()

        #polgridmore (17 columns)
        self.polgridmore = np.empty((0, 17))
        
        #polar pollutant concnetration (21 columns)
        self.polar_pol_con = np.empty((0, 21))
        
        #polar risk (28 columns)
        self.polar_risk = np.empty((0, 28))
        
        
        #apply polar build function and fill each matrix
        np.apply_along_axis(lambda x: self.polar_build(x), axis=1, arr=self.polgrid )


        # construct acute breakdown file object and write to file
        # acuteBkdn = AcuteBreakdown(self.outdir, self.facid, self.model, plot_df)
        # acuteBkdn.write()

        excelWriter = ExcelWriter(self.outdir)

        #writeout polgridmore
        headers_pg = ['utme','utmn','result','elev','hill','flag','avg_time','source_id',	'num_yrs','net_id',
              'distance', 'angle', 'utmz', 'lon', 'lat', 'sector', 'ring' ]
        excelWriter.write('polgridmore.xlsx', headers_pg, self.polgridmore)
        
        #write out polar pollutant concentration
        headers_pp = ['utme','utmn','result','elev','hill','flag','avg_time','source_id','num_yrs','net_id',
              'distance', 'angle', 'utmz', 'lon', 'lat', 'sector', 'ring', 'fac_id', 'pollutant', 'emis_tpy', 'conc']
        excelWriter.write('polar_pollutant_concentration.xlsx', headers_pp, self.polar_pol_con)
        
        #write our polar risk
        headers_pr = ['utme','utmn','result','elev','hill','flag','avg_time','source_id','num_yrs','net_id',
              'distance', 'angle', 'utmz', 'lon', 'lat', 'sector', 'ring', 'fac_id', 'pollutant', 'emis_tpy', 'conc',
              'ure', 'ufc', 'resp', 'neuro', 'risk', 'resp_hi', 'neur_hi']
        excelWriter.write('polar_risk.xlsx', headers_pr, self.polar_risk)

        # polar concentration (18 columns)
        self.pol_conc = np.empty((0,18))

        #create final risk matrix to store all risk computations
        self.finalrisk = np.empty((0, 35))

        #apply inner build function to get polar concentrations and risk computations
        np.apply_along_axis(self.inner_build, axis=1, arr=self.inner_m )

        #write out pollutant concentration
        headers_pc =['utme', 'utmn', 'result', 'elev', 'hill', 'flag', 'avg_time', 'source_id', 'num_yrs', 'net_id',
             'IDMARPLOT', 'LAT', 'LON', 'POPULATION', 'fac_id', 'pollutant', 'emis_tpy', 'conc']
        excelWriter.write('pollutant_concentration.xlsx', headers_pc, self.pol_conc)

        #write out inner risk
        headers_ir = ['utme', 'utmn', 'result', 'elev', 'hill', 'flag', 'avg_time', 'source_id', 'num_yrs', 'net_id',
              'IDMARPLOT', 'LAT', 'LON', 'POPULATION', 'fac_id', 'pollutant', 'emis_tpy', 'conc', 'ure', 'rfc', 'resp',
              'neuro', 'liver', 'dev', 'reprod', 'kidney', 'ocular', 'endoc', 'hemato', 'immune', 'skeletal', 'spleen',
              'thyroid', 'wholebod', 'risk', 'resp_hi', 'live_hi', 'neur_hi', 'deve_hi', 'repr_hi', 'kidn_hi', 'ocul_hi',
              'endo_hi', 'hema_hi', 'immu_hi', 'skel_hi', 'sple_hi', 'thyr_hi', 'whol_hi']
        excelWriter.write('inner_risk.xlsx', headers_ir, self.finalrisk)

#        MAXRISK TBD
#        #empty excel for max_individal_risk
#        self.max_risk = openpyxl.Workbook()
#        
#        #grab the active sheet
#        self.mr = self.max_risk.active
#        
#        #add column headers
#        self.mr.append(['utme', 'utmn', 'elev', 'hill', 'IDMARPLOT', 'LAT', 'LON', 'fac_id', 'risk', 'resp_hi', 'live_hi', 'neur_hi', 'dev_hi', 'repr_hi', 'kidn_hi', 'ocul_hi', 'endoc_hi', 'hem_hi', 'immu_hi', 'skel_hi', 'sple_hi', 'thyr_hi', 'whol_hi'])
        
        
        #loop through indices for each target endpoint
#        for i in range(20,35):
#            maxr = self.finalrisk[self.finalrisk[:,i].argmax()]
#            utm = maxr[0:2] #utme, utmn
#            elh = np.append(utm, maxr[3:5]) #elev, hill
#            ill = np.append(elh, maxr[10:13]) #idmarplot, lat, lon
#            fid = np.append(ill, maxr[14])
#            his = np.append(fid, maxr[20:])
#                
#            self.mr.append(his.tolist())



        #save to file
#        self.max_risk.save('max_risk.xlsx')
        
        

        #------------------ Outer Receptor Processing --------------------------------------

  
        
        
        #self.outer = np.empty((0,8))
    
        #empty excel for storing all outer receptors
        all_outer = openpyxl.Workbook()
        
        # grab the active worksheet
        ws = all_outer.active
        
        ws.append(['LAT', 'LON', 'conc_ug_m3', 'source_id', 'pollutant', 'population', 'FIPS', 'block', 'elev']) 
        
        pollutants = list(set(self.polar_pol_con[:, 18]))
        
        #convert to matrix
        matrix = self.outerblks_df.as_matrix() 
        
    
           
        
        for row in matrix:
            #get pollutant list based on sector and ring coordinates
            
            
            block = row[3][6:]
            
            
            s = ((row[14] * self.numsectors)/360.0 % self.numsectors) + 1
            if int(s) == self.numsectors:
                s1 = self.numsectors
                s2 = 1
            else:
                s1 = int(s)
                s2 = int(s) + 1
            r1 = int(row[19])
            if r1 == self.numrings:
                r1 = r1 - 1
            r2 = int(row[19]) + 1
            if r2 > self.numrings:
               r2 = self.numrings
            

            s1r1 = np.logical_and(self.polar_pol_con[:, 15] == s1, self.polar_pol_con[:,16] == r1 )
            s1r2 = np.logical_and(self.polar_pol_con[:, 15] == s1, self.polar_pol_con[:,16] == r2 )
            s2r1 = np.logical_and(self.polar_pol_con[:, 15] == s2, self.polar_pol_con[:,16] == r1 )
            s2r2 = np.logical_and(self.polar_pol_con[:, 15] == s2, self.polar_pol_con[:,16] == r2 )
            
            co1 = self.polar_pol_con[s1r1]
            co2 = self.polar_pol_con[s1r2]
            co3 = self.polar_pol_con[s2r1]
            co4 = self.polar_pol_con[s2r2]
            
            full = np.concatenate((co1, co2, co3, co4), axis=0)
            
            sourceid = full[:,7][0]
            
            for p in pollutants:
            
                sub = full[full[:,18] == p]
                conc = sub[:, 20]
                
                ring_loc = row[19]
                cs = row[17]
                
                if conc[0] == 0 or conc[1] == 0:
                    R_s12 = max(conc[0], conc[1])
                else:
                    Lnr_s12 = (math.log(conc[0]) * (int(ring_loc)+1-ring_loc)) + (math.log(conc[1]) * (ring_loc-int(ring_loc)))
                    R_s12 = math.exp(Lnr_s12)
            
                if conc[2] == 0 or conc[3] == 0:
                    R_s34 = max(conc[2], conc[3] )
                else:
                    Lnr_s34 = (math.log(conc[2]) * (int(ring_loc)+1-ring_loc)) + (math.log(conc[3] ) * (ring_loc-int(ring_loc)))
                    R_s34 = math.exp(Lnr_s34)
                cintconc = R_s12*(int(cs)+1-cs) + R_s34*(cs-int(cs))
                
                #write directly to excel, don't save anything in memory
                #self.outer = np.append(self.outer, [row[4], row[5], cintconc, sourceid, p, row[7], str(row[1]), str(block), row[0]], axis = 0)
                ws.append([row[4], row[5], cintconc, sourceid, p, row[7], str(row[1]), str(block), row[0]])
        
#          #create outerblocks excel FIX TO WORK WITH XLSX WRITER
#        outer = xlsxwriter.Workbook(self.outdir + 'all_outer.xlsx', {'constant_memory': True})
#        w_outer= outer.add_worksheet()
#        
#        headers_outer = ['LAT', 'LON', 'conc_ug_m3', 'source_id', 'pollutant', 'population', 'FIPS', 'block', 'elev']
#
#        
#        for o in range(0, len(headers_outer)):
#            ir.write(0, o, headers_outer[o]) 
#        
#        
#        for row6 in range(0, self.outer.shape[0]):
#            for col6 in range(0, 8):
#                w_outer.write(row6 + 1, col6, self.outer[row6][col6])
#                
#        outer.close
         
        all_outer.save(self.outdir + "all_outer_test.xlsx")
        print("Wrote all outer blocks")
        print(time.time()-start, 'seconds.')
        local_vars = inspect.currentframe().f_locals    
    
       
        return local_vars
    
    
    