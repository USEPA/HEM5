# -*- coding: utf-8 -*-
"""
Created on Mon Nov  3 15:26:04 2025

@author: SteveFudge
"""

import pandas as pd
from com.sca.hem4.log.Logger import Logger
import os
import shutil
import stat
import glob
from com.sca.hem4.writer.kml.KMLWriter_for_Merger import KMLWriter_for_Merger
import traceback
import sys
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from copy import copy


class MergeHemRuns():

    def __init__(self, orig_rundir, new_rundir):

        self.orig_rundir = orig_rundir
        self.new_rundir = new_rundir
        self.orig_rungroup_name = os.path.basename(self.orig_rundir)
        self.new_rungroup_name = os.path.basename(self.new_rundir)

        # Get the list of facilities used in the original and rerun. These will be used to filter the inputs.
        orig_faclist_file = os.path.join(self.orig_rundir, "Inputs", "faclist.xlsx")
        rerun_faclist_file = os.path.join(self.new_rundir, "Inputs", "faclist.xlsx")
        if os.path.exists(orig_faclist_file):
            orig_facs_df = pd.read_excel(orig_faclist_file, usecols=[0])
            self.orig_facs_list = orig_facs_df.iloc[:, 0].tolist()
        else:
            raise ValueError("File "+orig_faclist_file+" does not exist")
        if os.path.exists(rerun_faclist_file):
            rerun_facs_df = pd.read_excel(rerun_faclist_file, usecols=[0])
            self.rerun_facs_list = rerun_facs_df.iloc[:, 0].tolist()
        else:
            raise ValueError("File "+rerun_faclist_file+" does not exist")

    def make_writable_and_retry(self, func, path, exc_info):
        """
        Error handler for shutil.rmtree.
        If the error is access-related, attempts to change the file
        permissions to allow writing, then retries the removal operation.
        """
        # Check if the error is an Access Denied error
        if issubclass(exc_info[0], PermissionError):
            try:
                # Change the file attribute to be writable by the owner
                os.chmod(path, stat.S_IWUSR)
                # Retry the removal function
                func(path)
            except Exception as e:
                raise ValueError(f"Failed to change permissions and delete {path}: {e}")
        else:
            raise

        
        
    def PerformMerge(self):

        # 1. Delete any summary files, demographic assessment results, and kmz files
        
        # delete pop dir if it exists
        ejdir = os.path.join(self.orig_rundir, 'pop')        
        if os.path.exists(ejdir):
            try:
                shutil.rmtree(ejdir, onerror=self.make_writable_and_retry)
                Logger.logMessage("pop subdirectory deleted from the original HEM rungroup folder.\n")
            except OSError as e:
                Logger.logMessage("Error while trying to delete the pop subdirectory. Error message:\n"
                                  + e.strerror)
                raise ValueError(e.strerror)
        
        # delete Acute Maps dir if it exists        
        ejdir = os.path.join(self.orig_rundir, 'Acute Maps')        
        if os.path.exists(ejdir):
            try:
                shutil.rmtree(ejdir, onerror=self.make_writable_and_retry)
                Logger.logMessage("Acute Maps subdirectory deleted from the original HEM rungroup folder.\n")
            except OSError as e:
                Logger.logMessage("Error while trying to delete the Acute Maps subdirectory. Error message:\n"
                                  + e.strerror)
                raise ValueError(e.strerror)
        
        # delete kmz
        kmz_to_delete = glob.glob(os.path.join(self.orig_rundir, '*.kmz'))
        for file_path in kmz_to_delete:
            try:
                os.remove(file_path)
                Logger.logMessage(f"Deleted KMZ file {file_path}\n")
            except OSError as e:
                Logger.logMessage(f"Error trying to delete KMZ file {file_path}\n: {e}")
                raise ValueError(e.strerror)

        # delete summary files
        strings_to_exclude = ['facility_cancer_risk_exp','facility_max_risk_and_hi'
                              ,'facility_toshi_exp', 'hem.log']
        files_to_delete = []
        try:
            
            for filename in os.listdir(self.orig_rundir):
                if os.path.isfile(os.path.join(self.orig_rundir, filename)):  
                    exclude_file = False
                    for s in strings_to_exclude:
                        if s in filename:
                            exclude_file = True
                            break  # No need to check other strings if one is found
                    if not exclude_file:
                        files_to_delete.append(filename)
                        
            for fname in files_to_delete:
                fpath = os.path.join(self.orig_rundir, fname)
                if os.path.exists(fpath):
                    os.remove(fpath)
                        
        except OSError as e:
            Logger.logMessage("Error trying to delete summary files. Error message is:\n"
                              + e.strerror)
            raise ValueError(e.strerror)
        
  

        # 2. Update the original "facility" summary files

        # update one file at a time
        search_strings = ['facility_cancer_risk_exp', 'facility_max_risk_and_hi'
                            , 'facility_toshi_exp']
        
        for fstring in search_strings:
            orig_summary = os.path.join(self.orig_rundir, self.orig_rungroup_name+'_'+fstring+'.xlsx')
            new_summary = os.path.join(self.new_rundir, self.new_rungroup_name+'_'+fstring+'.xlsx')
            
            if os.path.exists(orig_summary):
                if os.path.exists(new_summary):
                    df = self.update_input(orig_summary, new_summary, 'A', 0)
                    os.remove(orig_summary)
                    df.to_excel(orig_summary, index=False)
                    self.copy_top_rows_with_formatting(new_summary, orig_summary, 0)
                    Logger.logMessage('Updated facility summary file \n' 
                                      + orig_summary + '\n')                   
                else:
                    Logger.logMessage("Error! Trying to update facility summary file\n"
                                      +new_summary+"\n"
                                      +"but file does not exist.\n")
                    raise ValueError("Error trying to update a facility summary file that does not exist.")
            else:
                Logger.logMessage("Error! Trying to update facility summary file\n"
                                  +orig_summary+"\n"
                                  +"but file does not exist.\n")
                raise ValueError("Error trying to update a facility summary file that does not exist.")
                    
                    
        
        # 3. Identify all common facilities in the two rungroups, delete them from the original,
        #    and copy the rerun facility folders to the original rungroup
        common_facs = self.find_common_subfolders(self.orig_rundir, self.new_rundir)
                
        if len(common_facs) > 0:
            
            for f in common_facs:
                origdir = os.path.join(self.orig_rundir, f)
                newdir = os.path.join(self.new_rundir, f)

                # delete original facility folder that was rerun
                if os.path.exists(origdir):
                    try:
                        shutil.rmtree(origdir, onerror=self.make_writable_and_retry)
                        Logger.logMessage(f"Directory '{origdir}' and its contents deleted successfully.\n")
                    except OSError as e:
                        Logger.logMessage(f"Error trying to delete directory {origdir}\nError is:\n"
                                          + f"{e.strerror}")
                        raise ValueError(e.strerror)
                else:
                    Logger.logMessage(f"Attemping to delete facility folder '{origdir}' but it does not exist.")
                    raise ValueError("Facility folder "+origdir+" does not exist.")
                    
                # copy facility folder from rerun group to original rungroup
                try:
                    # Copy the entire directory tree
                    shutil.copytree(newdir, origdir)
                    Logger.logMessage(f"Facility folder '{newdir}' \n"
                                      + f"successfully copied to '{origdir}'\n")
                except FileExistsError:
                    Logger.logMessage(f"Error. Destination folder '{origdir}' already exists. Please delete it or choose a different path.")
                except OSError as e:
                    Logger.logMessage("An error occurred while trying to copy a facility folder "+
                                      " from the rerun group to the original rungroup. Error is:\n"+
                                      e.strerror)
                    raise ValueError(e.strerror)
        
        
        # 4. Copy all facility folders from the rerun group that are not in the original rungroup
                
        new_folders = self.find_unique_subfolders(self.new_rundir, self.orig_rundir)
        
        if len(new_folders) > 0:
            for newfolder in new_folders:
                newfolder_src = os.path.join(self.new_rundir, newfolder)
                newfolder_dest = os.path.join(self.orig_rundir, newfolder)
                if os.path.exists(newfolder_src):
                    try:
                        shutil.copytree(newfolder_src, newfolder_dest)
                        Logger.logMessage(f"Facility folder '{newfolder_src}' \n"
                                          + f"successfully copied to '{newfolder_dest}'\n")
                    except OSError as e:
                        Logger.logMessage("An error occurred while trying to copy a facility folder "+
                                          " from the rerun group to the original rungroup. Error is:\n"+
                                          e.strerror)
                        raise ValueError(e.strerror)

        
        # 5. Update the files in the Inputs folder
        
        input_files = ['building_dimensions.xlsx'
                       ,'buoyant_line_parameters.xlsx'
                       ,'emisloc.xlsx'
                       ,'emisvar.xlsx'
                       ,'faclist.xlsx'
                       ,'hapemis.xlsx'
                       ,'landuse.xlsx'
                       ,'month-to-seasons.xlsx'
                       ,'particle_data.xlsx'
                       ,'polygon_vertex.xlsx'
                       ,'user_receptors.xlsx'
                       ,'haplib.xlsx'
                       ,'target_organs.xlsx']
        
        input_keys = [['A', 'B', 'C', 'D']
                      ,['A', 'B', 'C']
                      ,['A', 'B']
                      ,['A', 'B']
                      ,['A']
                      ,['A', 'B', 'C']
                      ,['A']
                      ,['A']
                      ,['A', 'B']
                      ,['A', 'B']
                      ,['A']
                      ,['A']
                      ,['A']]
        
        header_rows = [0
                       ,0
                       ,1
                       ,0
                       ,1
                       ,0
                       ,0
                       ,0
                       ,0
                       ,0
                       ,0
                       ,0
                       ,0]
        
        sheet_names = ['building dimensions'
                       ,'buoyant line'
                       ,'Emissions_Location'
                       ,'temporal season'
                       ,'Facility List Options'
                       ,'Hap emissions'
                       ,'sheet1'
                       ,'A'
                       ,'A'
                       ,'polygon vertex'
                       ,'user receptors'
                       ]
        
        try:

            # Initialize some dataframes needed for the KMZ
            faclist_df = None
            buoyant_df = None
            emisloc_df = None
            polygon_df = None
            
            
            for ifile, ikey, iheader, isheet in zip(input_files, input_keys, header_rows, sheet_names):
                orig_file = os.path.join(self.orig_rundir, "Inputs", ifile)
                new_file = os.path.join(self.new_rundir, "Inputs", ifile)
                                    
                # If input is only in rerun folder, filter to modeled facs, and copy to original folder
                if not os.path.exists(orig_file):
                    if os.path.exists(new_file):
                        self.copy_rerun_input(orig_file, new_file, iheader)
                        Logger.logMessage('Copied input file ' + ifile + ' to the original rungroup Inputs folder\n')
                    
                # Update if in both original and new. Also retain original header and sheetname.
                if os.path.exists(orig_file) and os.path.exists(new_file):
                    df = self.update_input(orig_file, new_file, ikey, iheader)
                    os.remove(orig_file)
                    df.to_excel(orig_file, index=False)
                    self.copy_top_rows_with_formatting(new_file, orig_file, iheader)
                    Logger.logMessage('Updated input file ' + ifile + '\n')
    
                # Store DFs for certain inputs because they'll be needed to create the KMZ
                if ifile == 'faclist.xlsx' and os.path.exists(orig_file):
                    faclist_df = pd.read_excel(orig_file, header=iheader, names=['fac_id','met_station'
                                ,'rural_urban','urban_pop','max_dist','model_dist','radial','circles'
                                ,'overlap_dist', 'ring1','fac_center','ring_distances', 'acute'
                                ,'hours','multiplier','hivalu','dep','depl','pdep','pdepl','vdep'
                                ,'vdepl','elev','flagpole','leadYN','user_rcpt','bldg_dw','fastall'
                                ,'emis_var','annual','period_start','period_end']
                                , dtype={'fac_id':str})
                elif ifile == 'buoyant_line_parameters.xlsx' and os.path.exists(orig_file):
                    buoyant_df = pd.read_excel(orig_file, header=iheader, names=['fac_id', 'blpgrp_id'
                                ,'source_id','avgbld_len','avgbld_hgt','avgbld_wid','avglin_wid'
                                ,'avgbld_sep','avgbuoy']
                                , dtype={'fac_id':str})
                elif ifile == 'emisloc.xlsx' and os.path.exists(orig_file):
                    emisloc_df = pd.read_excel(orig_file, header=iheader, names=['fac_id','source_id'
                                ,'location_type','lon','lat','utmzone','source_type','lengthx','lengthy'
                                ,'angle','horzdim','vertdim','areavolrelhgt','stkht','stkdia','stkvel'
                                ,'stktemp','elev','x2','y2','method','massfrac','partdiam']
                                , dtype={'fac_id':str})
                elif ifile == 'polygon_vertex.xlsx' and os.path.exists(orig_file):
                    polygon_df = pd.read_excel(orig_file, header=iheader, names=['fac_id','source_id'
                                ,'location_type','lon','lat','utmzone','numvert','area','fipstct']
                                , dtype={'fac_id':str})
                
        except BaseException as e:
            exc_type, exc_value, _ = sys.exc_info()
            formatted_exception = traceback.format_exception_only(exc_type, exc_value)
            Logger.logMessage('\nError occurred while tyring to update Input file ' + ifile 
                              + ' Error is:\n'
                              + formatted_exception[0])
            raise ValueError(formatted_exception[0])            


        # 6. Copy the rerun run log into the original rungroup and change the name of the log
        
        # rungroup = os.path.basename(self.orig_rundir)
        # old_name = os.path.join(self.orig_rundir, "hem.log")
        # new_name = os.path.join(self.orig_rundir, "hem-"+rungroup+".log")
        # os.rename(old_name, new_name)

        old_name = os.path.join(self.new_rundir, "hem.log")
        new_name = os.path.join(self.orig_rundir, "hem-"+self.new_rungroup_name+".log")
        shutil.copy(old_name, new_name)

        Logger.logMessage('Copied log file\n')
        

        # 7. Recreate the emission source KMZ
                
        kmlwriter = KMLWriter_for_Merger(self.orig_rundir, faclist_df, buoyant_df, emisloc_df, polygon_df)
        kmlwriter.write_kml_emis_loc()
        Logger.logMessage('Recreated the rungroup emission source KMZ file.\n')
        
        Logger.logMessage('\nMerging of HEM runs is complete.')
        


    def copy_rerun_input(self, dest_file, source_file, headerrow):
        
        # load source file into DF and filter to modeled facilities
        df = pd.read_excel(source_file, skiprows=headerrow, dtype=str)
        df = df[df.iloc[:,0].isin(self.rerun_faclist_file)]
        
        # export to destination Excel file
        df.to_excel(dest_file, index=False)
        
        # keep header and sheetname from the source Excel file
        self.copy_top_rows_with_formatting(source_file, dest_file, headerrow)
        return
        
        
        
    def update_input(self, orig_file, new_file, keycols, headerrow):
        
        # new is updating original
        
        df_source = pd.read_excel(new_file, skiprows=headerrow, dtype=str)
        # filter to modeled facilities
        df_source = df_source[df_source.iloc[:,0].isin(self.rerun_facs_list)]
        df_source = df_source.rename(columns=lambda y: self.conv(df_source.columns.get_loc(y)))
        
        df_target = pd.read_excel(orig_file, skiprows=headerrow, dtype=str)
        # filter to modeled facilities
        df_target = df_target[df_target.iloc[:,0].isin(self.orig_facs_list)]
        df_target = df_target.rename(columns=lambda y: self.conv(df_target.columns.get_loc(y)))
        
        # find common rows
        merged_df = df_target.merge(df_source, on=keycols, how='left', indicator=True
                                    , suffixes=['', '_right'])
        
        # remove common rows from target
        target_filtered = merged_df[merged_df['_merge'] == 'left_only'].drop(columns=['_merge'])
        
        # drop '_right' column names
        cols_to_drop = [col for col in target_filtered.columns if col.endswith('_right')]
        target_filtered = target_filtered.drop(cols_to_drop, axis=1)
        
        # Append source DF to filtered target.This will bring in any updated rows and new rows.
        target_final = pd.concat([target_filtered, df_source], ignore_index=True)
        target_final = target_final.sort_values(by=keycols, ignore_index=True)
        
        return target_final
        

    def conv(self, num):
        convStr = "ABCDEFGHIJKLMNOPQRSTUVWXYZ" # Assign any base you'd like
        b = len(convStr)
        if num<b:
            return convStr[num]
        else:
            return convStr[num//b-1] + convStr[num%b]
        


    def find_common_subfolders(self, folder1, folder2):
        """
        Finds all common subfolders between two given folders.
    
        Args:
            folder1 (str): The path to the first folder.
            folder2 (str): The path to the second folder.
    
        Returns:
            list: A list of relative paths to the common subfolders.
        """
 
        subfolders1 = set()
        for root, dirs, _ in os.walk(folder1):
            for d in dirs:
                relative_path = os.path.relpath(os.path.join(root, d), folder1)
                subfolders1.add(relative_path)
    
        subfolders2 = set()
        for root, dirs, _ in os.walk(folder2):
            for d in dirs:
                relative_path = os.path.relpath(os.path.join(root, d), folder2)
                subfolders2.add(relative_path)
    
        common_subfolders = list(subfolders1.intersection(subfolders2))
        
        # Remove the Inputs subfolder
        if 'Inputs' in common_subfolders:
            common_subfolders.remove('Inputs')
        
        return common_subfolders
    
    
    def find_unique_subfolders(self, dir1, dir2):
        """
        Finds subfolder names in dir1 that are not present in dir2.
    
        Args:
            dir1 (str): The path to the first directory.
            dir2 (str): The path to the second directory.
    
        Returns:
            list: A list of subfolder names unique to dir1.
        """
        # Get all immediate subdirectories in dir1
        subfolders_dir1 = {
            d for d in os.listdir(dir1) if os.path.isdir(os.path.join(dir1, d))
        }
    
        # Get all immediate subdirectories in dir2
        subfolders_dir2 = {
            d for d in os.listdir(dir2) if os.path.isdir(os.path.join(dir2, d))
        }
    
        # Find subfolders in dir1 that are not in dir2
        unique_subfolders = list(subfolders_dir1 - subfolders_dir2)
        return unique_subfolders
            

    def copy_top_rows_with_formatting(self, source_file, dest_file, numrows):
        """
        Copies the top n rows (including formatting) from a source Excel worksheet
        to the top of a destination Excel worksheet.
        """
        # Load workbooks
        source_wb = load_workbook(source_file)
        dest_wb = load_workbook(dest_file)
    
        # Get worksheets
        source_ws = source_wb.active
        dest_ws = dest_wb.active
        # source_ws = source_wb[source_sheet_name]
        # dest_ws = dest_wb[dest_sheet_name]
    
        # Insert n new rows at the top of the destination worksheet
        dest_ws.insert_rows(1, amount=numrows+1)
    
        # Iterate through the first n rows of the source worksheet
        for row_idx in range(1, numrows+2):  # Rows 1-n
            for col_idx in range(1, source_ws.max_column + 1):
                source_cell = source_ws.cell(row=row_idx, column=col_idx)
                dest_cell = dest_ws.cell(row=row_idx, column=col_idx)
    
                # Copy value
                dest_cell.value = source_cell.value
    
                # Copy formatting (font, border, fill, number_format, protection, alignment)
                if source_cell.has_style:
                    dest_cell.font = copy(source_cell.font)
                    dest_cell.border = copy(source_cell.border)
                    dest_cell.fill = copy(source_cell.fill)
                    dest_cell.number_format = copy(source_cell.number_format)
                    dest_cell.protection = copy(source_cell.protection)
                    dest_cell.alignment = copy(source_cell.alignment)

        # Copy merged cell ranges
        for merged_cell_range in source_ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_cell_range))
            
            # Only copy merged cells within the first n rows
            if min_row <= numrows+2 and max_row <= numrows+2:
                # Adjust the merged cell range to the new location (which is still rows 1 and 2)
                dest_ws.merge_cells(start_row=min_row, start_column=min_col,
                                    end_row=max_row, end_column=max_col)

                    
        # Delete row with temporary header from destination worksheet
        dest_ws.delete_rows(numrows+2)
        
    
        # Set the worksheet name in the destination workbook
        dest_ws.title = source_ws.title
        
        # Save the modified destination workbook
        dest_wb.save(dest_file)
        print(f"Top '{numrows}' rows copied from '{source_file}' to '{dest_file}' with formatting.")
