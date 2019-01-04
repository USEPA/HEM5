import xlsxwriter
from openpyxl import load_workbook
import pandas as pd

from writer.Writer import Writer

class ExcelWriter(Writer):

    def __init__(self, model, plot_df):
        Writer.__init__(self)

        self.model = model
        self.plot_df = plot_df

    def appendToFile(self, dataframe):
        data = dataframe.values
        book = load_workbook(self.filename)
        writer = pd.ExcelWriter(self.filename, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        ws = writer.book["Sheet1"]
        startrow = ws.max_row
        for row in range(0, data.shape[0]):
            for col in range(0, data.shape[1]):
                value = self.data[row][col]
                truncated = float('{:6.12}'.format(value)) if isinstance(value, float) else value
                ws.cell(row=startrow + row+1, column=col+1).value = truncated

        writer.save()

    def writeHeader(self):
        """
         Write the header (column names) in the given Excel worksheet.
        """
        workbook = xlsxwriter.Workbook(self.filename, {'constant_memory': True})
        worksheet = workbook.add_worksheet()

        headers = self.getHeader()
        for i in range(0, len(headers)):
            worksheet.write(0, i, headers[i])

        workbook.close()

    def analyze(self, data):
        pass